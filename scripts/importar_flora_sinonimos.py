"""
importar_flora_sinonimos.py
===========================
Lê angiospermsdatabase.xlsx e gymnospermsdatabase.xlsx,
extrai sinônimos com ocorrência no Cerrado e gera
flora_sinonimos_import.sql para importar no phpMyAdmin.

Dependências: pip install openpyxl
Uso:          python scripts/importar_flora_sinonimos.py
"""

import openpyxl
import os
import re
from datetime import datetime

# ── Configuração ──────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQUIVOS  = [
    os.path.join(BASE_DIR, 'angiospermsdatabase.xlsx'),
    os.path.join(BASE_DIR, 'gymnospermsdatabase.xlsx'),
]
SAIDA_SQL = os.path.join(BASE_DIR, 'database', 'flora_sinonimos_import.sql')
LOTE      = 500

# ── Índices das colunas ────────────────────────────────────────
COL_RANK        = 1
COL_GRUPO       = 2
COL_FAMILIA     = 6
COL_GENERO      = 9
COL_EPITETO     = 10
COL_AUTOR       = 14
COL_STATUS      = 16
COL_DOM         = 28
COL_SINONIMO_DE = 31

# ── Passo 1: coleta nomes aceitos do Cerrado ──────────────────
# (os mesmos critérios do importar_flora_brasil.py)
nomes_aceitos_cerrado = set()
for caminho in ARQUIVOS:
    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb['Relatório']
    for row in ws.iter_rows(min_row=2, values_only=True):
        rank   = str(row[COL_RANK]   or '').strip()
        status = str(row[COL_STATUS] or '').strip()
        dom    = str(row[COL_DOM]    or '')
        if rank == 'Espécie' and status == 'Nome aceito' and 'Cerrado' in dom:
            genero  = str(row[COL_GENERO]  or '').strip()
            epiteto = str(row[COL_EPITETO] or '').strip()
            nomes_aceitos_cerrado.add(f'{genero} {epiteto}'.strip())

print(f'Nomes aceitos do Cerrado carregados: {len(nomes_aceitos_cerrado)}')

# Prefixos de tipo de sinonímia a remover
PREFIXOS = ['heterotípico', 'homotípico', 'basônimo', 'basiônimo',
            'heterotipico', 'homotipico', 'basinimo']

# ── Helpers ───────────────────────────────────────────────────
def esc(v):
    if v is None or str(v).strip() == '':
        return 'NULL'
    s = str(v).strip()
    s = s.replace('\\', '\\\\').replace("'", "\\'")
    return f"'{s}'"

def extrair_tipo_e_nome(texto):
    """
    Recebe 'heterotípico Justicia congrua' e retorna ('heterotipico', 'Justicia congrua')
    Recebe 'Justicia congrua' e retorna ('', 'Justicia congrua')
    """
    texto = texto.strip()
    for pref in PREFIXOS:
        if texto.lower().startswith(pref.lower()):
            nome = texto[len(pref):].strip()
            tipo = pref.lower().replace('í','i').replace('ô','o').replace('â','a')
            return tipo, nome
    return '', texto

# ── Leitura ───────────────────────────────────────────────────
registros = []

for caminho in ARQUIVOS:
    nome_arq = os.path.basename(caminho)
    print(f'Lendo {nome_arq}...')

    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb['Relatório']

    aceitos = 0
    ignorados = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        rank       = str(row[COL_RANK]   or '').strip()
        status     = str(row[COL_STATUS] or '').strip()
        dom        = str(row[COL_DOM]    or '')
        sinonimo_de = row[COL_SINONIMO_DE]

        if rank != 'Espécie' or status != 'Sinônimo':
            ignorados += 1
            continue
        if not sinonimo_de:
            ignorados += 1
            continue

        genero  = str(row[COL_GENERO]  or '').strip()
        epiteto = str(row[COL_EPITETO] or '').strip()
        sinonimo_nome = f'{genero} {epiteto}'.strip()
        autor   = str(row[COL_AUTOR]   or '').strip() or None
        familia = str(row[COL_FAMILIA] or '').strip() or None

        # Pode haver múltiplos nomes aceitos separados por ';'
        partes = [p.strip() for p in str(sinonimo_de).split(';') if p.strip()]

        for parte in partes:
            tipo, nome_aceito = extrair_tipo_e_nome(parte)
            # Só importa se o nome aceito existe na nossa base do Cerrado
            if nome_aceito and nome_aceito in nomes_aceitos_cerrado:
                registros.append({
                    'sinonimo':    sinonimo_nome,
                    'autor':       autor,
                    'familia':     familia,
                    'nome_aceito': nome_aceito,
                    'tipo':        tipo or None,
                })
                aceitos += 1

    print(f'  Aceitos: {aceitos} | Ignorados: {ignorados}')

print(f'\nTotal de registros a importar: {len(registros)}')

# ── Geração do SQL ────────────────────────────────────────────
print(f'Gerando {SAIDA_SQL}...')

with open(SAIDA_SQL, 'w', encoding='utf-8') as f:
    f.write('-- ============================================================\n')
    f.write('-- flora_sinonimos_import.sql\n')
    f.write(f'-- Gerado em: {datetime.now().strftime("%Y-%m-%d %H:%M")}\n')
    f.write(f'-- Registros: {len(registros)} sinonimos do Cerrado\n')
    f.write('-- Fonte: REFLORA/JBRJ — CC-BY\n')
    f.write('-- ============================================================\n\n')
    f.write('SET NAMES utf8mb4;\n\n')

    cols = '`sinonimo`, `autor`, `familia`, `nome_aceito`, `tipo`'

    for i in range(0, len(registros), LOTE):
        lote = registros[i:i + LOTE]
        f.write(f'INSERT INTO `flora_brasil_sinonimos` ({cols}) VALUES\n')
        linhas = []
        for r in lote:
            linhas.append(
                f"({esc(r['sinonimo'])}, {esc(r['autor'])}, "
                f"{esc(r['familia'])}, {esc(r['nome_aceito'])}, {esc(r['tipo'])})"
            )
        f.write(',\n'.join(linhas))
        f.write(';\n\n')

print('Concluido.')
tamanho_kb = os.path.getsize(SAIDA_SQL) / 1024
print(f'Tamanho: {tamanho_kb:.0f} KB — OK para phpMyAdmin.')
